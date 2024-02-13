from openpyxl import Workbook
from openpyxl.styles import Font, Border, Side
import string
import tkinter as tk
from tkinter import Tk
from tkinter.filedialog import askopenfilename
from openpyxl import load_workbook

thick_top = Border(left=Side(style='thin'), 
                     right=Side(style='thin'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thin'))

thick = Border(left=Side(style='thick'), 
                     right=Side(style='thick'), 
                     top=Side(style='thick'), 
                     bottom=Side(style='thick'))


YTDWage = 0
YTDTax = 0
YTDCpp = 0
YTDEi = 0

wb = 0
Second = False

def makeHorizLine(ws, start, end, row):
    start = string.ascii_uppercase.index(start)
    end = string.ascii_uppercase.index(end)

    for letter in string.ascii_uppercase[start:end]:
        location = letter+str(row)
        ws[location].border = thick_top

def boldCell(ws, col, row):
    location = col+str(row)
    ws[location].border = thick

def getPrevFile():
    global YTDWage, YTDTax, YTDCpp, YTDEi

    Tk().withdraw() # we don't want a full GUI, so keep the root window from appearing
    filename = askopenfilename() # show an "Open" dialog box and return the path to the selected file
    wb_obj = load_workbook(filename)
    ws = wb_obj.active
    
    wage = 0
    tax = 0
    cpp = 0
    ei = 0

    if None != ws["H29"].value:
        wage = float(ws["H29"].value)
    elif None != ws["H9"]:
        wage = float(ws["H9"].value)

    if None != ws["H30"].value:
        tax = float(ws["H30"].value)
    elif None != ws["H10"].value:
        tax = float(ws["H10"].value)

    if None != ws["H31"].value:
        cpp = float(ws["H31"].value)
    elif None != ws["H11"].value:
        cpp = float(ws["H11"].value)

    if None != ws["H32"].value:
        ei = float(ws["H32"].value)
    elif None != ws["H12"].value:
        ei = float(ws["H12"].value)

    YTDWage = wage
    YTDTax = tax
    YTDCpp = cpp
    YTDEi = ei


    




def createGUI():
    # Create the main application window
    global wb

    def createFile():

        makeFile(c_entry.get(),
                p_entry.get(),
                pp_entry.get(),
                pa_entry.get(),
                vp_entry.get(),
                tax_entry.get(),
                cpp_entry.get(),
                ei_entry.get(),
                ws
                 )

        pp_entry.delete(0, 'end')
        pa_entry.delete(0, 'end')
        vp_entry.delete(0, 'end')
        tax_entry.delete(0, 'end')
        cpp_entry.delete(0, 'end')
        ei_entry.delete(0, 'end')

    # Create the main application window
    app = tk.Tk()
    app.title("Simple Paystub App")

    # Create a label
    title_label = tk.Label(app, text="Paystub App", font=("Helvetica", 16))
    title_label.pack(padx=500)

    wb = Workbook()
    ws = wb.active

    # Create an entry field
    c_label = tk.Label(app, text="company name")
    c_label.pack()
    c_entry = tk.Entry(app)
    c_entry.pack(pady=5)
    c_entry.config(width=c_entry["width"] + 40)

    p_label = tk.Label(app, text="person name")
    p_label.pack()
    p_entry = tk.Entry(app)
    p_entry.pack(pady=5)
    p_entry.config(width=p_entry["width"] + 40)

    pp_label = tk.Label(app, text="pay period")
    pp_label.pack()
    pp_entry = tk.Entry(app)
    pp_entry.pack(pady=5)

    pa_label = tk.Label(app, text="pay amount")
    pa_label.pack()
    pa_entry = tk.Entry(app)
    pa_entry.pack(pady=5)

    vp_label = tk.Label(app, text="vacation pay")
    vp_label.pack()
    vp_entry = tk.Entry(app)
    vp_entry.pack(pady=5)

    tax_label = tk.Label(app, text="tax")
    tax_label.pack()
    tax_entry = tk.Entry(app)
    tax_entry.pack(pady=5)

    cpp_label = tk.Label(app, text="CPP")
    cpp_label.pack()
    cpp_entry = tk.Entry(app)
    cpp_entry.pack(pady=5)

    ei_label = tk.Label(app, text="EI")
    ei_label.pack()
    ei_entry = tk.Entry(app)
    ei_entry.pack(pady=5)

    # Create a button
    create_button = tk.Button(app, text="Create File", command=createFile)
    create_button.pack(pady=5)

    prev_button = tk.Button(app, text="Get Previous File", command=getPrevFile)
    prev_button.pack(pady=5)

    # Create another label to display the greeting
    greeting_label = tk.Label(app, text="")
    greeting_label.pack(pady=5)

    
    # Run the application
    app.mainloop()
def checkDone(ws, filename):
    popup_window = tk.Toplevel()
    popup_window.title("Finished with worksheet?")

    def oneMore():
        global Second
        popup_window.destroy() 
        Second = True

    def finish(): 
        global Second, wb
        popup_window.destroy()
        Second = False
        wb.save(filename)

    label = tk.Label(popup_window, text="Are you finished?")
    label.pack(padx=100, pady=100)
    button = tk.Button(popup_window, text="One more", command=oneMore)
    button.pack(pady=10)
    button = tk.Button(popup_window, text="I'm done", command=finish)
    button.pack(pady=10)

        



def boldText(ws, col, row):
    location = col+str(row)
    ws[location].font = Font(bold=True)

def makeFile(c,p,pp,pa,vp,tax,cpp,ei,ws):
    global YTDWage, YTDTax, YTDCpp, YTDEi, Second
    company_name = c
    person_name = p
    pay_period = pp
    pay_amount = float(pa)
    vacation_pay = float(vp)
    tax = float(tax)
    cpp = float(cpp)
    ei = float(ei)

    gross_wage = pay_amount + vacation_pay
    net_wage = gross_wage - (cpp+ei+tax)
    baseline = 0

    page_title = "./sheets/" + person_name + pay_period + ".xlsx"
    if Second:
        baseline = 21
    else:
        baseline = 0

    line1 = baseline + 1
    line3 = baseline + 3
    line6 = baseline + 6
    line8 = baseline + 8
    line9 = baseline + 9
    line10 = baseline + 10
    line11 = baseline + 11
    line12 = baseline + 12
    line13 = baseline + 13
    line14 = baseline + 14
    line15 = baseline + 15
    line16 = baseline + 16
    line9 = baseline + 9
    line9 = baseline + 9


    makeHorizLine(ws, "F", "J", line9)
    makeHorizLine(ws, "F", "J", line13)
    makeHorizLine(ws, "F", "J", line14)
    makeHorizLine(ws, "F", "J", line15)
    makeHorizLine(ws, "F", "J", line16)

    boldCell(ws, "C", line9)
    boldCell(ws, "C", line10)

    writeFile(ws, "A", line1, "Company Name")
    writeFile(ws, "C", line1, company_name)
    boldText(ws, "A", line1)
    boldText(ws, "C", line1)

    writeFile(ws, "A", line3, "Pay To")
    writeFile(ws, "C", line3, person_name)

    writeFile(ws, "A", line6, "For The Pay Period")
    writeFile(ws, "C", line6, pay_period)

    writeFile(ws, "A", line9, "Pay")
    writeFile(ws, "C", line9, pay_amount)

    writeFile(ws, "A", line10, "Vacation Pay")
    writeFile(ws, "C", line10, vacation_pay)

    writeFile(ws, "E", line9, "Gross Wage")
    writeFile(ws, "F", line9, gross_wage)

    writeFile(ws, "E", line10, "Tax")
    writeFile(ws, "F", line10, tax)

    writeFile(ws, "E", line11, "CPP")
    writeFile(ws, "F", line11, cpp)
    writeFile(ws, "E", line12, "EI")
    writeFile(ws, "F", line12, ei)

    writeFile(ws, "E", line13, "Net Wage")
    writeFile(ws, "F", line13, net_wage)

    writeFile(ws, "E", line15, "Net Amount")
    writeFile(ws, "F", line15, net_wage)

    writeFile(ws, "F", line8, "Current")
    writeFile(ws, "H", line8, "Year To Date")

    YTDWage += gross_wage
    YTDTax += tax
    YTDCpp += cpp 
    YTDEi += ei

    writeFile(ws, "H", line9, YTDWage)
    writeFile(ws, "H", line10, YTDTax)
    writeFile(ws, "H", line11, YTDCpp) 
    writeFile(ws, "H", line12, YTDEi)

    writeFile(ws, "H", line13, YTDWage - (YTDCpp + YTDEi + YTDTax))
    writeFile(ws, "H", line15, YTDWage - (YTDCpp + YTDEi + YTDTax))
    

    page_title = "./sheets/" + person_name + pay_period + ".xlsx"

    if Second:
        global wb
        Second = False
        wb.save(page_title)
    else:
        checkDone(ws, page_title)





def writeFile(ws,col,row, value):
    location = col+str(row)
    ws[location] = value
    

def main():
    createGUI()




main()